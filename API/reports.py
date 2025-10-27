"""
Module de génération de rapports Excel et PDF
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas

from django.db.models import Sum, Count, F, Q
from django.utils import timezone

from .models import Produit, Vente, LigneVente, Achat, StockMove, ProductStock, Warehouse


class ExcelReportGenerator:
    """Générateur de rapports Excel avec styles professionnels"""

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

        # Styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def set_column_widths(self, widths):
        """Définir les largeurs de colonnes"""
        for idx, width in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(idx)].width = width

    def add_title(self, title, row=1):
        """Ajouter un titre au rapport"""
        self.ws.merge_cells(f'A{row}:F{row}')
        cell = self.ws[f'A{row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        return row + 1

    def add_metadata(self, row, company_name="GestionStock", date=None):
        """Ajouter métadonnées du rapport"""
        if date is None:
            date = timezone.now().strftime('%d/%m/%Y %H:%M')

        self.ws[f'A{row}'] = "Date:"
        self.ws[f'B{row}'] = date
        self.ws[f'A{row+1}'] = "Société:"
        self.ws[f'B{row+1}'] = company_name
        return row + 3

    def add_header_row(self, headers, row):
        """Ajouter une ligne d'en-tête"""
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        return row + 1

    def add_data_rows(self, data, start_row, number_columns=None):
        """Ajouter des lignes de données"""
        if number_columns is None:
            number_columns = []

        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right' if col_idx in number_columns else 'left')

                # Format numbers
                if col_idx in number_columns and isinstance(value, (int, float, Decimal)):
                    if isinstance(value, Decimal) or isinstance(value, float):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'

        return row_idx + 1

    def add_total_row(self, row, label_col, label, value_cols_values):
        """Ajouter une ligne de total"""
        cell = self.ws.cell(row=row, column=label_col)
        cell.value = label
        cell.font = Font(bold=True)
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right')

        for col_idx, value in value_cols_values.items():
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='right')
            if isinstance(value, (int, float, Decimal)):
                if isinstance(value, Decimal) or isinstance(value, float):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'

        return row + 1

    def save_to_buffer(self):
        """Sauvegarder dans un buffer BytesIO"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer


class PDFReportGenerator:
    """Générateur de rapports PDF avec styles professionnels"""

    def __init__(self, title, orientation='portrait'):
        self.buffer = BytesIO()
        pagesize = landscape(A4) if orientation == 'landscape' else A4
        self.doc = SimpleDocTemplate(self.buffer, pagesize=pagesize,
                                     rightMargin=30, leftMargin=30,
                                     topMargin=30, bottomMargin=30)
        self.elements = []
        self.styles = getSampleStyleSheet()
        self.title = title

        # Styles personnalisés
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )

    def add_title(self, title=None):
        """Ajouter un titre"""
        if title is None:
            title = self.title
        self.elements.append(Paragraph(title, self.title_style))
        self.elements.append(Spacer(1, 12))

    def add_metadata(self, company_name="GestionStock", date=None):
        """Ajouter métadonnées"""
        if date is None:
            date = timezone.now().strftime('%d/%m/%Y %H:%M')

        metadata_style = ParagraphStyle(
            'Metadata',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        )

        self.elements.append(Paragraph(f"<b>Date:</b> {date}", metadata_style))
        self.elements.append(Paragraph(f"<b>Société:</b> {company_name}", metadata_style))
        self.elements.append(Spacer(1, 20))

    def add_table(self, data, col_widths=None):
        """Ajouter un tableau"""
        table = Table(data, colWidths=col_widths)

        # Style du tableau
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Corps du tableau
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Grille
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # Alternance de couleurs
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        self.elements.append(table)
        self.elements.append(Spacer(1, 20))

    def add_summary_box(self, summary_data):
        """Ajouter un encadré de résumé"""
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        self.elements.append(summary_table)
        self.elements.append(Spacer(1, 20))

    def build(self):
        """Construire le document PDF"""
        self.doc.build(self.elements)
        self.buffer.seek(0)
        return self.buffer


# ======================
# RAPPORTS SPÉCIFIQUES
# ======================

def generate_stock_valuation_excel(warehouse_id=None):
    """
    Rapport de valorisation du stock (Excel)
    Montre la valeur totale du stock par produit
    """
    gen = ExcelReportGenerator()
    gen.ws.title = "Valorisation Stock"

    # Titre
    current_row = gen.add_title("RAPPORT DE VALORISATION DU STOCK")
    current_row = gen.add_metadata(current_row)

    # Requête
    products = Produit.objects.filter(is_active=True).select_related('categorie', 'fournisseur', 'currency')

    if warehouse_id:
        # Stock par entrepôt
        warehouse = Warehouse.objects.get(pk=warehouse_id)
        gen.ws[f'A{current_row}'] = f"Entrepôt: {warehouse.name}"
        current_row += 2

        product_stocks = ProductStock.objects.filter(
            warehouse=warehouse,
            quantity__gt=0
        ).select_related('produit', 'produit__categorie')

        products = [ps.produit for ps in product_stocks]
    else:
        gen.ws[f'A{current_row}'] = "Tous les entrepôts"
        current_row += 2

    # En-têtes
    headers = ['Référence', 'Désignation', 'Catégorie', 'Qté Stock', 'Prix Unit.', 'Valeur Stock', 'Statut']
    gen.set_column_widths([15, 30, 20, 12, 15, 18, 15])
    current_row = gen.add_header_row(headers, current_row)

    # Données
    data_rows = []
    total_value = Decimal('0.00')
    total_qty = 0

    for product in products:
        qty = product.quantite
        if qty <= 0:
            continue

        unit_price = product.prixU or Decimal('0.00')
        value = qty * unit_price
        total_value += value
        total_qty += qty

        data_rows.append([
            product.reference,
            product.designation,
            product.categorie.nom if product.categorie else 'N/A',
            qty,
            float(unit_price),
            float(value),
            product.get_stock_status()
        ])

    current_row = gen.add_data_rows(data_rows, current_row, number_columns=[4, 5, 6])

    # Total
    gen.add_total_row(current_row, 3, "TOTAL", {
        4: total_qty,
        6: float(total_value)
    })

    return gen.save_to_buffer()


def generate_stock_valuation_pdf(warehouse_id=None):
    """
    Rapport de valorisation du stock (PDF)
    """
    gen = PDFReportGenerator("RAPPORT DE VALORISATION DU STOCK", orientation='landscape')
    gen.add_title()
    gen.add_metadata()

    # Requête
    products = Produit.objects.filter(is_active=True, quantite__gt=0).select_related('categorie', 'fournisseur')

    if warehouse_id:
        warehouse = Warehouse.objects.get(pk=warehouse_id)
        gen.elements.append(Paragraph(f"<b>Entrepôt:</b> {warehouse.name}", gen.styles['Normal']))
        gen.elements.append(Spacer(1, 12))

        product_stocks = ProductStock.objects.filter(
            warehouse=warehouse,
            quantity__gt=0
        ).select_related('produit')
        products = [ps.produit for ps in product_stocks]

    # Tableau
    table_data = [['Référence', 'Désignation', 'Catégorie', 'Qté', 'Prix Unit.', 'Valeur Stock', 'Statut']]

    total_value = Decimal('0.00')
    total_qty = 0

    for product in products:
        qty = product.quantite
        unit_price = product.prixU or Decimal('0.00')
        value = qty * unit_price
        total_value += value
        total_qty += qty

        table_data.append([
            product.reference,
            product.designation[:30],
            product.categorie.nom if product.categorie else 'N/A',
            str(qty),
            f"{unit_price:.2f}",
            f"{value:.2f}",
            product.get_stock_status()
        ])

    # Ligne de total
    table_data.append(['', '', 'TOTAL', str(total_qty), '', f"{total_value:.2f}", ''])

    gen.add_table(table_data, col_widths=[1.2*inch, 2.5*inch, 1.3*inch, 0.8*inch, 1*inch, 1.2*inch, 1*inch])

    # Résumé
    summary_data = [
        ['Nombre de produits en stock:', str(len(products))],
        ['Quantité totale:', str(total_qty)],
        ['Valeur totale du stock:', f"{total_value:.2f}"]
    ]
    gen.add_summary_box(summary_data)

    return gen.build()


def generate_sales_report_excel(start_date=None, end_date=None):
    """
    Rapport des ventes (Excel)
    """
    gen = ExcelReportGenerator()
    gen.ws.title = "Rapport Ventes"

    # Titre
    current_row = gen.add_title("RAPPORT DES VENTES")

    # Période
    if start_date and end_date:
        gen.ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = gen.ws[f'A{current_row}']
        cell.value = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        cell.alignment = Alignment(horizontal='center')
        current_row += 1

    current_row = gen.add_metadata(current_row)

    # Requête
    ventes = Vente.objects.filter(statut='completed').select_related('client', 'warehouse')

    if start_date:
        ventes = ventes.filter(date_vente__gte=start_date)
    if end_date:
        ventes = ventes.filter(date_vente__lte=end_date)

    ventes = ventes.order_by('-date_vente')

    # En-têtes
    headers = ['N° Vente', 'Date', 'Client', 'Type Paiement', 'HT', 'TTC', 'Remise %']
    gen.set_column_widths([15, 12, 25, 15, 15, 15, 12])
    current_row = gen.add_header_row(headers, current_row)

    # Données
    data_rows = []
    total_ht = Decimal('0.00')
    total_ttc = Decimal('0.00')

    for vente in ventes:
        total_ht += vente.total_ht or Decimal('0.00')
        total_ttc += vente.total_ttc or Decimal('0.00')

        data_rows.append([
            vente.numero,
            vente.date_vente.strftime('%d/%m/%Y'),
            f"{vente.client.nom} {vente.client.prenom}" if vente.client else 'N/A',
            vente.get_type_paiement_display(),
            float(vente.total_ht or 0),
            float(vente.total_ttc or 0),
            float(vente.remise_percent or 0)
        ])

    current_row = gen.add_data_rows(data_rows, current_row, number_columns=[5, 6, 7])

    # Total
    gen.add_total_row(current_row, 4, "TOTAL", {
        5: float(total_ht),
        6: float(total_ttc)
    })

    return gen.save_to_buffer()


def generate_sales_report_pdf(start_date=None, end_date=None):
    """
    Rapport des ventes (PDF)
    """
    gen = PDFReportGenerator("RAPPORT DES VENTES", orientation='landscape')
    gen.add_title()

    # Période
    if start_date and end_date:
        period_style = ParagraphStyle('Period', parent=gen.styles['Normal'], fontSize=11, alignment=1)
        gen.elements.append(Paragraph(
            f"<b>Période:</b> {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            period_style
        ))
        gen.elements.append(Spacer(1, 12))

    gen.add_metadata()

    # Requête
    ventes = Vente.objects.filter(statut='completed').select_related('client')

    if start_date:
        ventes = ventes.filter(date_vente__gte=start_date)
    if end_date:
        ventes = ventes.filter(date_vente__lte=end_date)

    ventes = ventes.order_by('-date_vente')

    # Tableau
    table_data = [['N° Vente', 'Date', 'Client', 'Type Paiement', 'Montant HT', 'Montant TTC', 'Remise %']]

    total_ht = Decimal('0.00')
    total_ttc = Decimal('0.00')
    count_ventes = 0

    for vente in ventes:
        total_ht += vente.total_ht or Decimal('0.00')
        total_ttc += vente.total_ttc or Decimal('0.00')
        count_ventes += 1

        table_data.append([
            vente.numero,
            vente.date_vente.strftime('%d/%m/%Y'),
            f"{vente.client.nom} {vente.client.prenom}"[:25] if vente.client else 'N/A',
            vente.get_type_paiement_display(),
            f"{vente.total_ht or 0:.2f}",
            f"{vente.total_ttc or 0:.2f}",
            f"{vente.remise_percent or 0:.1f}%"
        ])

    # Ligne de total
    table_data.append(['', '', '', 'TOTAL', f"{total_ht:.2f}", f"{total_ttc:.2f}", ''])

    gen.add_table(table_data, col_widths=[1.2*inch, 1*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 0.8*inch])

    # Résumé
    ticket_moyen = total_ttc / count_ventes if count_ventes > 0 else Decimal('0.00')
    summary_data = [
        ['Nombre de ventes:', str(count_ventes)],
        ['Chiffre d\'affaires HT:', f"{total_ht:.2f}"],
        ['Chiffre d\'affaires TTC:', f"{total_ttc:.2f}"],
        ['Ticket moyen:', f"{ticket_moyen:.2f}"]
    ]
    gen.add_summary_box(summary_data)

    return gen.build()


def generate_inventory_report_excel():
    """
    Rapport d'inventaire complet (Excel)
    """
    gen = ExcelReportGenerator()
    gen.ws.title = "Inventaire"

    # Titre
    current_row = gen.add_title("RAPPORT D'INVENTAIRE COMPLET")
    current_row = gen.add_metadata(current_row)

    # En-têtes
    headers = ['Référence', 'Code Barre', 'Désignation', 'Catégorie', 'Fournisseur', 'Qté', 'Seuil Alerte', 'Prix Unit.', 'Statut']
    gen.set_column_widths([15, 15, 30, 20, 20, 10, 12, 15, 15])
    current_row = gen.add_header_row(headers, current_row)

    # Données
    products = Produit.objects.filter(is_active=True).select_related('categorie', 'fournisseur').order_by('reference')

    data_rows = []
    for product in products:
        data_rows.append([
            product.reference,
            product.code_barre or '',
            product.designation,
            product.categorie.nom if product.categorie else '',
            product.fournisseur.libelle if product.fournisseur else '',
            product.quantite,
            product.seuil_alerte or 0,
            float(product.prixU or 0),
            product.get_stock_status()
        ])

    gen.add_data_rows(data_rows, current_row, number_columns=[6, 7, 8])

    return gen.save_to_buffer()


def generate_inventory_report_pdf():
    """
    Rapport d'inventaire complet (PDF)
    """
    gen = PDFReportGenerator("RAPPORT D'INVENTAIRE COMPLET", orientation='landscape')
    gen.add_title()
    gen.add_metadata()

    products = Produit.objects.filter(is_active=True).select_related('categorie', 'fournisseur').order_by('reference')

    # Tableau
    table_data = [['Réf.', 'Désignation', 'Catégorie', 'Fournisseur', 'Qté', 'Seuil', 'Prix', 'Statut']]

    for product in products:
        table_data.append([
            product.reference,
            product.designation[:25],
            product.categorie.nom[:15] if product.categorie else '',
            product.fournisseur.libelle[:15] if product.fournisseur else '',
            str(product.quantite),
            str(product.seuil_alerte or 0),
            f"{product.prixU or 0:.2f}",
            product.get_stock_status()
        ])

    gen.add_table(table_data, col_widths=[0.9*inch, 2*inch, 1.2*inch, 1.2*inch, 0.7*inch, 0.7*inch, 0.9*inch, 1*inch])

    # Statistiques
    rupture = products.filter(quantite=0).count()
    critique = products.filter(quantite__gt=0, quantite__lte=F('seuil_critique')).count()
    alerte = products.filter(quantite__gt=F('seuil_critique'), quantite__lte=F('seuil_alerte')).count()
    normal = products.filter(quantite__gt=F('seuil_alerte')).count()

    summary_data = [
        ['Produits en rupture:', str(rupture)],
        ['Produits en critique:', str(critique)],
        ['Produits en alerte:', str(alerte)],
        ['Produits normaux:', str(normal)],
        ['Total produits:', str(products.count())]
    ]
    gen.add_summary_box(summary_data)

    return gen.build()
